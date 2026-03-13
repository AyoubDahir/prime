import frappe
from frappe.utils import getdate
from openpyxl import load_workbook
import os
import calendar
from frappe.utils.file_manager import get_file_path


def import_employee_schedule(
    file_path: str,
    year: int,
    month: int,
    month_name: str = None,
    doctype: str = "Employee Schedulling",
    employee_columns=("Employee", "Employee ID", "EmployeeID", "Emp ID", "EmpID"),
    skip_columns=("Employee Name",),
    shift_map=None,
    commit: bool = True,
):
    """
    Import employee schedules from an Excel sheet where:
      - first row is headers
      - one column is employee id (Employee / Employee ID / ...)
      - remaining columns are day numbers (1..31) with values like D/N/DN/OFF

    Usage:
      bench execute prime.setup.data_importer.import_schedule_jan_2026
      OR
      bench execute 'prime.setup.data_importer.import_employee_schedule' --kwargs "{...}"
    """
    if shift_map is None:
        shift_map = {
            "D": "Day Shift",
            "N": "Night Shift",
            "DN": "Day and Night Shift",
            "ND": "Night Day Shift",
            "CANTEEN": "CANTEEN",
            "OFF": "Free",
            "OF": "Free",
        }

    # derive month label if not provided
    if not month_name:
        import calendar
        month_name = calendar.month_name[int(month)]

    wb = load_workbook(file_path)
    sheet = wb.active

    headers = [cell.value for cell in sheet[1]]

    # find employee id column
    emp_col = next((c for c in employee_columns if c in headers), None)
    if not emp_col:
        frappe.throw(f"Employee column not found. Found headers: {headers}")

    inserted = 0
    skipped = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))

        emp_id = row_dict.get(emp_col)
        if not emp_id:
            continue

        emp_name = frappe.db.get_value("Employee", emp_id, "employee_name")

        for key, value in row_dict.items():
            if key in (emp_col,) + tuple(skip_columns):
                continue
            if value is None:
                continue
            if not isinstance(value, str):
                continue

            code = value.strip().upper()
            shift = shift_map.get(code)
            if not shift:
                skipped += 1
                continue

            # day header might be 1 or "1"
            try:
                day_int = int(str(key).strip())
            except Exception:
                skipped += 1
                continue

            shift_date = f"{int(year):04d}-{int(month):02d}-{day_int:02d}"

            doc = frappe.get_doc({
                "doctype": doctype,
                "employee": emp_id,
                "employee_name": emp_name,
                "shift": shift,
                "from_date": getdate(shift_date),
                "to_date": getdate(shift_date),
                "day": str(day_int),
                "label": shift_date,
                "month": month_name,
                "year": str(year),
            })

            doc.insert()
            inserted += 1

    if commit:
        frappe.db.commit()

    return {"inserted": inserted, "skipped": skipped, "month": month_name, "year": year}


# ---- Monthly wrappers (so you don't pass kwargs every time) ----

def import_schedule_jan_2026():
    res = import_employee_schedule(
        file_path="/home/hussein/frappe-bench/royaljan.xlsx",
        year=2026,
        month=1,
        month_name="January",
    )
    print(res)
    return res


def import_schedule_dec_2025():
    res = import_employee_schedule(
        file_path="/home/hussein/frappe-bench/royal_decemebr.xlsx",
        year=2025,
        month=12,
        month_name="December",
    )
    print(res)
    return res


@frappe.whitelist()
def import_employee_schedule_from_file(file_url: str, year: int, month: int, overwrite: int = 0):
    """
    Import schedules from an uploaded File (file_url).
    - overwrite=0: skip if employee+date already exists
    - overwrite=1: update existing record's shift + dates

    Returns summary dict.
    """
    year = int(year)
    month = int(month)
    overwrite = int(overwrite or 0)

    if not file_url:
        frappe.throw("Missing file_url")

    # Resolve file path on disk
    path = get_file_path(file_url)
    if not os.path.exists(path):
        frappe.throw(f"File not found on disk: {path}")

    month_name = calendar.month_name[month]

    shift_map = {
        "D": "Day Shift",
        "N": "Night Shift",
        "DN": "Day and Night Shift",
        "ND": "Night Day Shift",
        "CANTEEN": "CANTEEN",
        "OFF": "Free",
        "OF": "Free",
    }

    wb = load_workbook(path)
    sheet = wb.active

    headers = [cell.value for cell in sheet[1]]

    # detect employee id column name (December vs January)
    employee_columns = ("Employee", "Employee ID", "EmployeeID", "Emp ID", "EmpID")
    emp_col = next((c for c in employee_columns if c in headers), None)
    if not emp_col:
        frappe.throw(f"Employee column not found. Found headers: {headers}")

    inserted = 0
    updated = 0
    skipped = 0
    errors = 0

    # Small performance boost: cache employee name
    emp_name_cache = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))

        emp_id = row_dict.get(emp_col)
        if not emp_id:
            continue

        if emp_id not in emp_name_cache:
            emp_name_cache[emp_id] = frappe.db.get_value("Employee", emp_id, "employee_name") or ""

        emp_name = emp_name_cache[emp_id]

        for key, value in row_dict.items():
            if key in (emp_col, "Employee Name"):
                continue
            if value is None:
                continue
            if not isinstance(value, str):
                continue

            code = value.strip().upper()
            shift = shift_map.get(code)
            if not shift:
                skipped += 1
                continue

            try:
                day_int = int(str(key).strip())
            except Exception:
                skipped += 1
                continue

            shift_date = f"{year:04d}-{month:02d}-{day_int:02d}"
            shift_date_d = getdate(shift_date)

            try:
                existing = frappe.db.get_value(
                    "Employee Schedulling",
                    {"employee": emp_id, "from_date": shift_date_d},
                    "name",
                )

                if existing:
                    if overwrite:
                        doc = frappe.get_doc("Employee Schedulling", existing)
                        doc.shift = shift
                        doc.employee_name = emp_name
                        doc.to_date = shift_date_d
                        doc.day = str(day_int)
                        doc.label = shift_date
                        doc.month = month_name
                        doc.year = str(year)
                        doc.save(ignore_permissions=True)
                        updated += 1
                    else:
                        skipped += 1
                    continue

                doc = frappe.get_doc({
                    "doctype": "Employee Schedulling",
                    "employee": emp_id,
                    "employee_name": emp_name,
                    "shift": shift,
                    "from_date": shift_date_d,
                    "to_date": shift_date_d,
                    "day": str(day_int),
                    "label": shift_date,
                    "month": month_name,
                    "year": str(year),
                })
                doc.insert(ignore_permissions=True)
                inserted += 1

            except Exception:
                errors += 1
                frappe.log_error(frappe.get_traceback(), "Schedule Import Row Error")

    frappe.db.commit()

    return {
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
        "month": month_name,
        "year": year,
        "file_url": file_url,
        "overwrite": overwrite,
    }


# import frappe
# import pandas as pd
# from frappe.utils import getdate

# def create_rooms():
#     df = pd.read_excel(r'/home/hussein/frappe-bench/royal_decemebr.xlsx')
#     df = pd.DataFrame(df)
#     data = df.to_dict(orient='records')
    
#     formatted_data = []
#     for item in data:
#         formatted_item = {'Employee': item['Employee']}
#         shifts = {key: value for key, value in item.items() if key not in ['Employee', 'Employee Name']}
#         formatted_item['shifts'] = shifts
#         formatted_data.append(formatted_item)

#     try:
#         for f_data in formatted_data:
#             emp_id = f_data['Employee']
#             emp_name = frappe.db.get_value("Employee", emp_id, "employee_name")
#             for key, value in f_data['shifts'].items():
#                 if not isinstance(value, str):
#                     continue
#                 value = value.strip().upper()
#                 shift = ""
#                 if value == "D":
#                     shift = "Day Shift"
#                 elif value == "N":
#                     shift = "Night Shift"
#                 elif value == "DN":
#                     shift = "Day and Night Shift"
#                 elif value == "ND":
#                     shift = "Night Day Shift"
#                 elif value == "CANTEEN":
#                     shift = "CANTEEN"
#                 elif value in ["OFF", "OF"]:
#                     shift = "Free"
                
#                 if not shift:
#                     continue

#                 try:
#                     day_int = int(key)
#                     shift_date = f"2025-12-{day_int:02d}"
#                 except ValueError:
#                     continue

#                 sched_doc = frappe.get_doc({
#                     "doctype": "Employee Schedulling",  # Double check spelling!
#                     "employee": emp_id,
#                     "employee_name": emp_name,
#                     "shift": shift,
#                     "from_date": getdate(shift_date),
#                     "to_date": getdate(shift_date),
#                     "day": key,
#                     "label": shift_date,
#                     "month": "December",
#                     "year": "2025"
#                 })
#                 sched_doc.insert()

#         frappe.db.commit()
#     except Exception as error:
#         frappe.log_error(frappe.get_traceback(), "Create Rooms Error")
#         print(error)
